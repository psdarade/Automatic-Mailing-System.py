from PIL import ImageTk;
from tkinter import *
import tkinter as tk
import tkinter.messagebox as tmsg
from tkinter.filedialog import askopenfile
import time
from tkinter.ttk import *
import smtplib,ssl
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import holidays
from datetime import datetime
import schedule



root = Tk()
root.title("Email_automation")
root.geometry("1000x700")
root.config(bg="grey")
photo=ImageTk.PhotoImage(file="bg.png")
l1=Label(image=photo)
l1.place(x=0,y=0)
canvas=Canvas(bg="darkorange",width=720,height=500)
canvas.place(x=120,y=100)

name=tk.Label(text="Email Address",bg="darkorange",font="Lucida 30")
name.place(x=150,y=210)
name=tk.Label(text="Password",bg="darkorange",font="Lucida 30")
name.place(x=150,y=310)

namevalue=StringVar()
namevalue1=StringVar()
nameEntry=Entry(root,textvariable=namevalue,show="*",font="Lucida 19",width=24)
nameEntry1=Entry(root,textvariable=namevalue1,font="Lucida 19",width=24)
nameEntry1.place(x=430,y=220)
nameEntry.place(x=430,y=320)
c=0;
def command():
    email=namevalue1.get();
    pas=namevalue.get();
    count=0;

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(email, pas);
    except:
        count=1;

    if count==0:
        newWindow = Toplevel(root)
        newWindow.geometry("900x600")
        newWindow.config(bg="#ccffff")
        var = StringVar()
        var.set(1)

        radio1 = tk.Radiobutton(newWindow, bg="#ccffff", text="Holiday", font="Lucida 15", variable=var,
                                value="Holiday")
        radio2 = tk.Radiobutton(newWindow, bg="#ccffff", text="Weekends", font="Lucida 15", variable=var,
                                value="Weekends")
        radio3 = tk.Radiobutton(newWindow, bg="#ccffff", text="Special Holiday", font="Lucida 15", variable=var,
                                value="Special Holiday")
        radio1.place(x=110, y=250)
        radio2.place(x=110, y=300)
        radio3.place(x=110, y=350)

        def selection():
            res = tmsg.askquestion(newWindow, message=f"Do you want to send mail for {var.get()}")
            if res == 'yes':
                try:
                    path=file_path.name;
                except:
                    tmsg.showerror("Error", "please select a file");
                    return
                selected=var.get()

                if selected==1:
                    tmsg.showerror("Error","select an option");
                    return

                c=1;
                present = datetime.now()
                h = holidays.India(years=2021);
                h1 = holidays.India(years=2022);
                dates = [];
                for date, ocasion in h.items():
                    if date >= present.date():
                        dates.append(date);
                for date, ocasion in h1.items():
                    dates.append(date);

                wb = openpyxl.load_workbook(path);
                sheet = wb.active;
                m_row = sheet.max_row;

                sender_email = email
                reciever_email = []
                password = pas

                for i in range(1, m_row):
                    cell = sheet.cell(row=i + 1, column=2)
                    reciever_email.append(cell.value);

                message = MIMEMultipart("alternative")
                message["Subject"] = "Annoncement"
                message["From"] = sender_email

                name = []
                for i in range(1, m_row):
                    cell = sheet.cell(row=i + 1, column=1)
                    name.append(cell.value);

                def holiday_mail():
                    count = 0;
                    for dt in dates:
                        if dt == present.date():
                            count = 1;
                    if count == 1:
                        for i in range(0, m_row - 1):
                            html = "<html><body><p>Hi {},<br>".format(name[i]);
                            html1 = html + "<br><center><h4>Seasons greetings, team!</h4></center><br>"\
                                           "As you may have noticed, the holiday season is around the corner." \
                                           " we hope that you and your family are safe and in happy holiday spirits!" \
                                           "<br><br>During this special time of year,we want to make sure you know how much" \
                                           " we appreciate you.Hope your holiday season is full of health" \
                                           " and happiness!<br><br>Thank you again for being the bet team one " \
                                           " could ask for.Your energy and persistence has driven our company to new " \
                                           " heights, and we greatly appreciate it.<br><br>Happy holidays," \
                                           "<br><br>Sincerely,<br><br>Management.</p></body></html>"
                            send = MIMEText(html1, "html")
                            message.attach(send)
                            context = ssl.create_default_context()
                            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                                server.login(sender_email, password)
                                server.sendmail(sender_email, reciever_email[i], message.as_string())
                        print('mail sent');
                    else:
                        print('not sent');

                def weekend_mail():
                    for i in range(0, m_row - 1):
                        html = "<html><body><p>Hi {},<br>".format(name[i]);
                        html1 = html + "<br><center><h4>Seasons greetings, team!</h4></center><br>"\
                                           "As you may have noticed, the weekend season is around the corner." \
                                           " we hope that you and your family are safe and in happy weekend spirits!" \
                                           "<br><br>During this special time of year,we want to make sure you know how much" \
                                           " we appreciate you.Hope your weekend season is full of health" \
                                           " and happiness!<br><br>Thank you again for being the best team one " \
                                           " could ask for.Your energy and persistence has driven our company to new " \
                                           " heights, and we greatly appreciate it.<br><br>Happy weekends," \
                                           "<br><br>Sincerely,<br><br>Management.</p></body></html>"
                        send = MIMEText(html1, "html")
                        message.attach(send)
                        context = ssl.create_default_context()
                        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                            server.login(sender_email, password)
                            server.sendmail(sender_email, reciever_email[i], message.as_string())

                def send_mail():
                    for i in range(0, m_row - 1):
                        html = "<html><body><p>Hi {},<br>".format(name[i]);
                        html1 = html + "<br><center><h4>Seasons greetings, team!</h4></center><br>"\
                                       "This mail is to inform you that your comapany has registered" \
                                       " your email in a system specially designed to automate emails" \
                                       " informing about weekends, Holidays and some of Special Holidays." \
                                       "<br><br>Thank You,<br>" \
                                       "<br>Sincerely,<br><br>Management.</p></body></html>"
                        send = MIMEText(html1, "html")
                        message.attach(send)
                        context = ssl.create_default_context()
                        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                            server.login(sender_email, password)
                            server.sendmail(sender_email, reciever_email[i], message.as_string())

                send_mail();
                if selected=="Holiday" or selected=="Special Holiday":
                    schedule.every().day.at("10:00").do(holiday_mail);
                elif selected=="Weekends":
                    schedule.every().saturday.at("10:00").do(weekend_mail);
                tmsg.showinfo("Success","process completed sucessfully");

            else:
                tmsg.showinfo('Return', 'Returnning to main appliction')

        def open_file():
            global file_path;
            file_path = askopenfile(mode='r', filetypes=[('email', '*xlsx')])
            if file_path is not None:
                pass

        def uploadfiles():
            try:
                print(file_path);
                pb1 = Progressbar(newWindow, orient=HORIZONTAL, length=520,
                                  mode='determinate')
                pb1.place(x=110, y=150)
                for i in range(5):
                    newWindow.update_idletasks()
                    pb1['value'] += 20
                    time.sleep(1)
                pb1.destroy()
                tk.Label(newWindow, bg="#ccffff", font="Bold 13", text='File Uploaded Successfully!',
                         foreground='black').place(x=360, y=200)
            except:
                tmsg.showerror("Error","please select a file");

        adhar = tk.Label(
            newWindow, font="Bold 15", bg="#ccffff",
            text='Upload file in excel format ')
        adhar.place(x=110, y=80)

        adharbtn = tk.Button(
            newWindow,
            text='Choose File', font="Bold 10", bg="white",
            command=lambda:open_file())
        adharbtn.place(x=650, y=80)

        upld = tk.Button(newWindow, bg="white", text='Upload Files', font="Bold 10", command=uploadfiles)
        upld.place(x=650, y=150)

        b1 = tk.Button(newWindow, bg="white", text="Send", font="Bold 15", command=selection)
        b1.place(x=410, y=460)
    else:
        tmsg.showerror("Error","Invalid email credentials");


buttonimg=ImageTk.PhotoImage(file="login.png")
button = tk.Button(root,image=buttonimg, command=command)
button.place(x=400, y=450)

root.mainloop()

if c==1:
    while True:
        schedule.run_pending()
        time.sleep(1);


